import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
import pyxlsb
import os
import re
import tempfile
from datetime import datetime
from dateutil.relativedelta import relativedelta
from io import BytesIO


# ================================================================
# Data Reading
# ================================================================

def read_ahmed_forecast(path_or_buf):
    """Read 'Final Supplies Fcst SKU to Base' tab.
    Returns dict[(region, sku)] = [18 forecast values]
    """
    data = {}
    with pyxlsb.open_workbook(path_or_buf) as wb:
        with wb.get_sheet("Final Supplies Fcst SKU to Base") as ws:
            for i, row in enumerate(ws.rows()):
                vals = [c.v for c in row]
                if i == 0:
                    continue
                sku = vals[4]   # Primary Base Product
                reg = vals[5]   # Theater
                if not sku or not reg:
                    continue
                fcst = [v if v is not None else 0 for v in vals[8:26]]
                data[(reg, sku)] = fcst
    return data


def read_ahmed_actuals(path_or_buf, col_letter):
    """Read 'L10 SHIP BASE ex-BRAZIL' tab actuals column.
    Returns dict[(region, sku)] = actual_value
    """
    col_idx = ord(col_letter.upper()) - ord("A")
    data = {}
    with pyxlsb.open_workbook(path_or_buf) as wb:
        with wb.get_sheet("L10 SHIP BASE ex-BRAZIL") as ws:
            for i, row in enumerate(ws.rows()):
                if i == 0:
                    continue
                vals = [c.v for c in row]
                sku = vals[3]   # Primary Base Product
                reg = vals[4]   # Theater
                if not sku or not reg:
                    continue
                val = vals[col_idx] if col_idx < len(vals) else None
                data[(reg, sku)] = val if val is not None else 0
    return data


def read_master(path_or_buf):
    """Read Master Vlookup 'Table' tab.
    Lookup key = Material (Col A). Returns dict[material] = {...}
    """
    lookup = {}
    wb = openpyxl.load_workbook(path_or_buf, data_only=True)
    ws = wb["Table"]
    for r in range(3, ws.max_row + 1):
        material = ws.cell(r, 1).value   # Col A = Material
        if not material:
            continue
        key = str(material).strip()
        lookup[key] = {
            "platform":  ws.cell(r, 3).value,    # Col C = Product Name
            "factor":    ws.cell(r, 6).value,    # Col F = Unit Quantity
            "plc":       ws.cell(r, 7).value,    # Col G = PLC
            "canon_cap": ws.cell(r, 13).value,   # Col M = Canon-Shared Cap
            "pl":        ws.cell(r, 18).value,   # Col R = PL
        }
    wb.close()
    return lookup


def _parse_sheet_header_date(value):
    """Parse either a datetime cell or a Feb-26 style header string."""
    if isinstance(value, datetime):
        return datetime(value.year, value.month, 1)
    if isinstance(value, str):
        text = value.strip()
        try:
            parsed = datetime.strptime(text, "%b-%y")
            return datetime(parsed.year, parsed.month, 1)
        except ValueError:
            return None
    return None


def read_prev_compare(path_or_buf, m1_label):
    """Read previous compare file, extract M-1 rows.
    Returns (dict[(reg,sku)] = {date: value}, set of available dates)
    """
    m1_data = {}
    wb = openpyxl.load_workbook(path_or_buf, data_only=True)
    ws = wb["Compare Packs"]

    # Map column numbers to dates from header
    date_cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        parsed_date = _parse_sheet_header_date(v)
        if parsed_date is not None:
            date_cols[c] = parsed_date
    available_dates = set(date_cols.values())

    # Read rows matching M-1 label
    for r in range(2, ws.max_row + 1):
        cycle = ws.cell(r, 8).value
        if cycle != m1_label:
            continue
        reg = ws.cell(r, 1).value
        sku = ws.cell(r, 2).value
        if not reg or not sku:
            continue
        reg = str(reg).strip()
        sku = str(sku).strip()
        vals = {}
        for c, dt in date_cols.items():
            v = ws.cell(r, c).value
            vals[dt] = v if v is not None else 0
        m1_data[(reg, sku)] = vals

    wb.close()
    return m1_data, available_dates


# ================================================================
# Date Generation
# ================================================================

def make_output_dates(m0_cycle):
    """Generate 19 output dates: 1 actuals month + 18 forecast months."""
    year = int(m0_cycle[:4])
    month = int(m0_cycle[4:6])

    # Actuals = month before M0
    act_y, act_m = (year, month - 1) if month > 1 else (year - 1, 12)
    dates = [datetime(act_y, act_m, 1)]

    # 18 forecast months starting from M0
    for i in range(18):
        m = month + i
        y = year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        dates.append(datetime(y, m, 1))

    return dates


# ================================================================
# Core Processing
# ================================================================

def process(ahmed_buf, prev_buf, master_buf,
            m0_cycle, m1_cycle, m0_label, m1_label, actuals_col):
    """Main processing: read all files, compute compare, return (bytes, count)."""

    fcst_data = read_ahmed_forecast(ahmed_buf)
    actuals_data = read_ahmed_actuals(ahmed_buf, actuals_col)
    master = read_master(master_buf)
    m1_data, prev_avail_dates = read_prev_compare(prev_buf, m1_label)

    output_dates = make_output_dates(m0_cycle)

    # Union of all (region, sku) keys from Ahmed forecast and previous compare
    all_keys = set(fcst_data.keys()) | set(m1_data.keys())

    factor_map = {}
    rows = []
    for reg, sku in sorted(all_keys):
        # Master lookup - skip if not found
        info = master.get(str(sku).strip())
        if not info:
            continue

        # Cache factor for Singles calculation
        f = info.get("factor")
        factor_map[str(sku).strip()] = f if isinstance(f, (int, float)) and f else 1

        conc = f"{reg}{sku}"
        base = [reg, sku, info["platform"], info["canon_cap"],
                info["pl"], info["plc"], conc]

        # --- M0 row: actuals + 18 forecast ---
        actual = actuals_data.get((reg, sku), 0)
        fcst = fcst_data.get((reg, sku), [0] * 18)
        m0_vals = [actual] + list(fcst)

        # --- M-1 row: align by date ---
        m1_dict = m1_data.get((reg, sku), None)
        m1_vals = []
        for dt in output_dates:
            if m1_dict is not None:
                m1_vals.append(m1_dict.get(dt))
            elif dt in prev_avail_dates:
                m1_vals.append(0)
            else:
                m1_vals.append(None)

        # --- DELTA row ---
        delta = []
        for i in range(len(output_dates)):
            v0 = m0_vals[i] if i < len(m0_vals) and m0_vals[i] is not None else 0
            v1 = m1_vals[i] if i < len(m1_vals) and m1_vals[i] is not None else 0
            delta.append(v0 - v1)

        rows.append(base + [m0_label] + m0_vals)
        rows.append(base + [m1_label] + m1_vals)
        rows.append(base + ["DELTA"] + delta)

    # Write to temp file (needed for win32com PivotTable creation)
    tmp_dir = tempfile.mkdtemp()
    tmp_path = os.path.join(tmp_dir, "output.xlsx")
    write_excel(rows, output_dates, m0_label, m1_label,
                m0_cycle, m1_cycle, tmp_path, factor_map)

    with open(tmp_path, "rb") as f:
        result_bytes = f.read()
    os.remove(tmp_path)
    os.rmdir(tmp_dir)

    return result_bytes, len(rows) // 3


# ================================================================
# Excel Output
# ================================================================

# ---------- Shared styles ----------

_HEADER_FILL = PatternFill(
    fgColor=Color(theme=8, tint=0.7999816888943144), fill_type="solid")
_HEADER_FONT = Font(bold=True)
_M1_FILL = PatternFill(
    fgColor=Color(theme=0, tint=-0.0499893185216834), fill_type="solid")
_DELTA_FILL = PatternFill(
    fgColor=Color(theme=7, tint=0.7999816888943144), fill_type="solid")
_GOLD_FILL = PatternFill(
    fgColor=Color(theme=7, tint=0.7999816888943144), fill_type="solid")
_ORANGE_FILL = PatternFill(
    fgColor=Color(theme=5, tint=0.7999816888943144), fill_type="solid")

_NUM_FMT = r'#,##0_ ;[Red]\-#,##0\ '
_PIVOT_NUM_FMT = r'#,##0_);[Red](#,##0)'


def _date_label(dt):
    """Format datetime as 'Feb-26' style label."""
    return dt.strftime("%b-%y")


def _make_singles_rows(rows, factor_map):
    """Multiply all value columns by the per-SKU factor."""
    singles = []
    for row_data in rows:
        sku = str(row_data[1]).strip()
        factor = factor_map.get(sku, 1)
        new_row = list(row_data[:8])  # base fields + cycle label unchanged
        for v in row_data[8:]:
            if isinstance(v, (int, float)):
                new_row.append(v * factor)
            else:
                new_row.append(v)
        singles.append(new_row)
    return singles


def write_excel(rows, dates, m0_label, m1_label,
                m0_cycle, m1_cycle, path, factor_map):
    """Write output workbook with 6 tabs matching reference format."""
    wb = openpyxl.Workbook()

    singles_rows = _make_singles_rows(rows, factor_map)

    # --- Packs sheets ---
    ws_pivot = wb.active
    ws_pivot.title = "Pivot Packs"
    ws_cbp = wb.create_sheet("Changes by Platform Packs")
    ws_cp = wb.create_sheet("Compare Packs")
    ws_pd = wb.create_sheet("PivotData")

    # --- Singles sheets ---
    ws_cs = wb.create_sheet("Compare Singles")
    ws_cbps = wb.create_sheet("Changes by Platform Singles")
    ws_pivots = wb.create_sheet("Pivot Singles")
    ws_pds = wb.create_sheet("PivotData Singles")

    _write_compare_packs(ws_cp, rows, dates, m0_label, m1_label)
    _write_changes_by_platform(ws_cbp, rows, dates, m0_label, m1_label,
                               m0_cycle, m1_cycle)
    _write_pivot_data(ws_pd, rows, dates)

    _write_compare_packs(ws_cs, singles_rows, dates, m0_label, m1_label)
    _write_changes_by_platform(ws_cbps, singles_rows, dates, m0_label, m1_label,
                               m0_cycle, m1_cycle, label="SINGLES")
    _write_pivot_data(ws_pds, singles_rows, dates)

    wb.save(path)

    # Post-process: create real PivotTables + CUMM Delta via Excel COM
    _create_pivot_tables(path, rows, singles_rows, dates, m0_cycle)


# ==================== Compare Packs ====================

def _write_compare_packs(ws, rows, dates, m0_label, m1_label):
    # Header row
    date_labels = [_date_label(dt) for dt in dates]
    header = ["Reg", "PRODUCT_MODEL_NR", "PRODUCT_PLATFORM_NM",
              "Canon-Shared Cap", "PL", "PLC", "ConC",
              "Forecast Cycle"] + date_labels
    for c, v in enumerate(header, 1):
        cell = ws.cell(1, c, v)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    # Data rows
    for r_idx, row_data in enumerate(rows, 2):
        cycle = row_data[7]
        for c_idx, v in enumerate(row_data, 1):
            cell = ws.cell(r_idx, c_idx, v)
            # Row fill
            if cycle == m1_label:
                cell.fill = _M1_FILL
            elif cycle == "DELTA":
                cell.fill = _DELTA_FILL
            # Number format for date columns
            if c_idx >= 9:
                cell.number_format = _NUM_FMT

    # Autofilter
    last_col = get_column_letter(len(header))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # Freeze panes (just below header, to the right of col H)
    ws.freeze_panes = "I2"


# ==================== Changes by Platform ====================

def _write_changes_by_platform(ws, rows, dates, m0_label, m1_label,
                               m0_cycle, m1_cycle, label="PACKS"):
    # Determine month short names from cycles
    m0_month = datetime(int(m0_cycle[:4]), int(m0_cycle[4:6]), 1).strftime("%b")
    m1_month = datetime(int(m1_cycle[:4]), int(m1_cycle[4:6]), 1).strftime("%b")

    # Use first 12 dates, split into 6 + 6
    first6 = dates[0:6]
    second6 = dates[6:12]

    # Aggregate DELTA by platform
    plat_delta = {}
    for row_data in rows:
        if row_data[7] != "DELTA":
            continue
        plat = row_data[2]  # PRODUCT_PLATFORM_NM
        vals = row_data[8:]  # all 19 date values
        if plat not in plat_delta:
            plat_delta[plat] = [0.0] * len(dates)
        for i, v in enumerate(vals):
            if isinstance(v, (int, float)):
                plat_delta[plat][i] += v

    # Sort by sum of first 6 months descending (most positive first)
    # Secondary sort: alphabetical (case-insensitive) for ties
    sorted_plats = sorted(
        plat_delta.items(),
        key=lambda x: (-sum(x[1][:6]), x[0].upper())
    )

    # --- Row 1: label + SUBTOTAL formulas ---
    ws.cell(1, 1, label).font = Font(bold=True)
    # SUBTOTAL for first 6 date cols (B-G)
    for ci in range(2, 8):
        col_l = get_column_letter(ci)
        cell = ws.cell(1, ci, f"=SUBTOTAL(9,{col_l}4:{col_l}9999)")
        cell.number_format = _NUM_FMT
    # SUBTOTAL for second 6 date cols (H-M)
    for ci in range(8, 8 + len(second6)):
        col_l = get_column_letter(ci)
        cell = ws.cell(1, ci, f"=SUBTOTAL(9,{col_l}4:{col_l}9999)")
        cell.number_format = _NUM_FMT
    # SUBTOTAL for remaining empty cols (N-W) to match reference
    for ci in range(8 + len(second6), 24):
        col_l = get_column_letter(ci)
        cell = ws.cell(1, ci, f"=SUBTOTAL(9,{col_l}4:{col_l}9999)")
        cell.number_format = _NUM_FMT

    # --- Row 2: Title ---
    ws.cell(2, 1, f"{m0_month} vs {m1_month} Fcst Delta").font = Font(bold=True)

    # --- Row 3: Headers ---
    ws.cell(3, 1, "Row Labels")
    ws.cell(3, 1).fill = _HEADER_FILL
    ws.cell(3, 1).font = _HEADER_FONT

    # First 6 date columns (B-G)
    for i, dt in enumerate(first6):
        ci = 2 + i
        cell = ws.cell(3, ci, _date_label(dt))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    # Second 6 date columns (H-M)
    for i, dt in enumerate(second6):
        ci = 8 + i
        cell = ws.cell(3, ci, _date_label(dt))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    # --- Row 4+: Platform data ---
    for r_off, (plat, all_vals) in enumerate(sorted_plats):
        r = 4 + r_off
        ws.cell(r, 1, plat)

        # First 6 months
        for i in range(6):
            cell = ws.cell(r, 2 + i, all_vals[i])
            cell.number_format = _NUM_FMT

        # Second 6 months
        for i in range(len(second6)):
            cell = ws.cell(r, 8 + i, all_vals[6 + i])
            cell.number_format = _NUM_FMT

    # Autofilter: A3:W{last_row} matching reference
    last_data_row = 3 + len(sorted_plats)
    ws.auto_filter.ref = f"A3:W{last_data_row}"

    # Freeze panes
    ws.freeze_panes = "B4"


# ==================== PivotData (hidden source sheet) ====================

def _write_pivot_data(ws, rows, dates):
    """Write flat DELTA-only data for PivotTable source.
    Columns: Reg, PRODUCT_MODEL_NR, PRODUCT_PLATFORM_NM, Forecast Cycle,
             then one column per month with string header like 'Feb-26'.
    """
    month_labels = [_date_label(dt) for dt in dates]
    header = ["Reg", "PRODUCT_MODEL_NR", "PRODUCT_PLATFORM_NM",
              "Forecast Cycle"] + month_labels
    for c, v in enumerate(header, 1):
        ws.cell(1, c, v)

    r = 2
    for row_data in rows:
        if row_data[7] != "DELTA":
            continue
        ws.cell(r, 1, row_data[0])   # Reg
        ws.cell(r, 2, row_data[1])   # PRODUCT_MODEL_NR
        ws.cell(r, 3, row_data[2])   # PRODUCT_PLATFORM_NM
        ws.cell(r, 4, row_data[7])   # Forecast Cycle = "DELTA"
        for ci, v in enumerate(row_data[8:]):
            ws.cell(r, 5 + ci, v if isinstance(v, (int, float)) else 0)
        r += 1

    ws.sheet_state = 'hidden'


# ==================== PivotTable + CUMM Delta via win32com ====================

def _build_one_pivot(wb, data_sheet, pivot_sheet, a1_label, table_name,
                     rows, dates):
    """Create one real PivotTable + CUMM Delta section on the given sheets."""
    # Pre-compute CUMM Delta data
    delta_dates = dates[:18]
    act = dates[0]
    cumm_start = act - relativedelta(months=1)
    cumm_dates = [cumm_start] + list(delta_dates)

    regions = ["AMERICAS", "ASIA PACIFIC", "EMEA"]
    region_delta = {r: [0.0] * len(dates) for r in regions}
    for row_data in rows:
        if row_data[7] != "DELTA":
            continue
        reg = row_data[0]
        if reg not in region_delta:
            continue
        for i, v in enumerate(row_data[8:]):
            if isinstance(v, (int, float)):
                region_delta[reg][i] += v

    grand = [0.0] * len(dates)
    for reg in regions:
        for i in range(len(dates)):
            grand[i] += region_delta[reg][i]

    ws_data = wb.Sheets(data_sheet)
    ws_pivot = wb.Sheets(pivot_sheet)

    # ---- Create PivotTable ----
    last_row = ws_data.Cells(ws_data.Rows.Count, 1).End(-4162).Row  # xlUp
    last_col = ws_data.Cells(1, ws_data.Columns.Count).End(-4159).Column  # xlToLeft
    data_range = ws_data.Range(
        ws_data.Cells(1, 1),
        ws_data.Cells(last_row, last_col)
    )

    pc = wb.PivotCaches().Create(
        SourceType=1,  # xlDatabase
        SourceData=data_range
    )

    pt = pc.CreatePivotTable(
        TableDestination=ws_pivot.Range('A7'),
        TableName=table_name
    )

    # Write label at A1 (above PivotTable)
    ws_pivot.Cells(1, 1).Value = a1_label
    ws_pivot.Cells(1, 1).Font.Bold = True

    # Filters (page fields) — add in reverse visual order
    pf_plat = pt.PivotFields('PRODUCT_PLATFORM_NM')
    pf_plat.Orientation = 3  # xlPageField

    pf_cycle = pt.PivotFields('Forecast Cycle')
    pf_cycle.Orientation = 3  # xlPageField
    pf_cycle.CurrentPage = 'DELTA'

    # Row field: Reg only
    pf_reg = pt.PivotFields('Reg')
    pf_reg.Orientation = 1  # xlRowField
    pf_reg.Position = 1

    # Value fields: Sum of each delta month column (18 months)
    month_labels = [_date_label(dt) for dt in delta_dates]
    for ml in month_labels:
        df = pt.AddDataField(
            pt.PivotFields(ml),
            'Sum of %s' % ml,
            -4157  # xlSum
        )
        df.NumberFormat = _PIVOT_NUM_FMT

    # Layout: no subtotals
    pf_reg.Subtotals = (False,) * 12

    pt.ColumnGrand = True
    pt.RowGrand = False

    # Show classic PivotTable layout with in-grid drop zones
    pt.InGridDropZones = True

    # Fix header label: "Row Labels" -> "Reg"
    pt.CompactLayoutRowHeader = 'Reg'

    # ---- Write CUMM Delta at fixed row 19 ----
    r19 = 19
    cell_cumm = ws_pivot.Cells(r19, 1)
    cell_cumm.Value = 'CUMM Delta'
    cell_cumm.Font.Bold = True

    gold_rgb = 0xC0DCF3  # BGR for Excel (light gold ~#F3DCC0)
    cell_cumm.Interior.Color = gold_rgb

    r20 = r19 + 1
    ws_pivot.Cells(r20, 1).Value = 'Region'
    ws_pivot.Cells(r20, 1).Interior.Color = gold_rgb
    cumm_labels = ['Sum of %s' % _date_label(dt) for dt in cumm_dates]
    for ci, lbl in enumerate(cumm_labels):
        c = ws_pivot.Cells(r20, 2 + ci)
        c.Value = lbl
        c.Interior.Color = gold_rgb

    # Region cumulative rows
    for ri, reg in enumerate(regions):
        cumm_r = r20 + 1 + ri
        ws_pivot.Cells(cumm_r, 1).Value = reg
        running = 0.0
        for ci in range(len(cumm_dates)):
            if ci == 0:
                running = region_delta[reg][0]
            elif ci < len(delta_dates):
                running += region_delta[reg][ci]
            cell = ws_pivot.Cells(cumm_r, 2 + ci)
            cell.Value = running
            cell.NumberFormat = _PIVOT_NUM_FMT

    # Grand Total cumulative row
    gt_row = r20 + 1 + len(regions)
    gt_rgb = 0xD5D5D5  # light gray for Grand Total
    ws_pivot.Cells(gt_row, 1).Value = 'Grand Total'
    ws_pivot.Cells(gt_row, 1).Font.Bold = True
    ws_pivot.Cells(gt_row, 1).Interior.Color = gt_rgb
    running = 0.0
    for ci in range(len(cumm_dates)):
        if ci == 0:
            running = grand[0]
        elif ci < len(delta_dates):
            running += grand[ci]
        cell = ws_pivot.Cells(gt_row, 2 + ci)
        cell.Value = running
        cell.NumberFormat = _PIVOT_NUM_FMT
        cell.Font.Bold = True
        cell.Interior.Color = gt_rgb

    # Column widths
    ws_pivot.Columns('A').ColumnWidth = 29.5
    ws_pivot.Columns('B').ColumnWidth = 15.125
    ws_pivot.Columns('H').ColumnWidth = 15.125
    ws_pivot.Columns('I').ColumnWidth = 15.125


def _calc_region_totals(rows, dates):
    """Compute DELTA totals by region for pivot/fallback output."""
    regions = ["AMERICAS", "ASIA PACIFIC", "EMEA"]
    region_delta = {r: [0.0] * len(dates) for r in regions}
    for row_data in rows:
        if row_data[7] != "DELTA":
            continue
        reg = row_data[0]
        if reg not in region_delta:
            continue
        for i, v in enumerate(row_data[8:]):
            if isinstance(v, (int, float)):
                region_delta[reg][i] += v

    grand = [0.0] * len(dates)
    for reg in regions:
        for i in range(len(dates)):
            grand[i] += region_delta[reg][i]

    return regions, region_delta, grand


def _write_one_pivot_fallback(ws, rows, dates, a1_label):
    """Fallback pivot writer (non-COM), used when win32com is unavailable."""
    delta_dates = dates[:18]
    cumm_dates = [dates[0] - relativedelta(months=1)] + list(delta_dates)
    regions, region_delta, grand = _calc_region_totals(rows, dates)

    ws.cell(1, 1, a1_label).font = Font(bold=True)
    ws.cell(4, 1, "Forecast Cycle")
    ws.cell(4, 2, "DELTA")
    ws.cell(5, 1, "PRODUCT_PLATFORM_NM")
    ws.cell(5, 2, "(All)")

    ws.cell(7, 2, "Values")
    ws.cell(8, 1, "Reg")
    for ci, dt in enumerate(delta_dates):
        ws.cell(8, 2 + ci, f"Sum of {_date_label(dt)}")

    for ri, reg in enumerate(regions):
        r = 9 + ri
        ws.cell(r, 1, reg)
        for ci in range(len(delta_dates)):
            c = ws.cell(r, 2 + ci, region_delta[reg][ci])
            c.number_format = _PIVOT_NUM_FMT

    gt_row = 9 + len(regions)
    ws.cell(gt_row, 1, "Grand Total").font = Font(bold=True)
    for ci in range(len(delta_dates)):
        c = ws.cell(gt_row, 2 + ci, grand[ci])
        c.number_format = _PIVOT_NUM_FMT
        c.font = Font(bold=True)

    r19 = 19
    ws.cell(r19, 1, "CUMM Delta").font = Font(bold=True)
    ws.cell(r20 := r19 + 1, 1, "Region")
    for ci, dt in enumerate(cumm_dates):
        ws.cell(r20, 2 + ci, f"Sum of {_date_label(dt)}")

    for ri, reg in enumerate(regions):
        r = r20 + 1 + ri
        ws.cell(r, 1, reg)
        running = 0.0
        for ci in range(len(cumm_dates)):
            if ci == 0:
                running = region_delta[reg][0]
            elif ci < len(delta_dates):
                running += region_delta[reg][ci]
            c = ws.cell(r, 2 + ci, running)
            c.number_format = _PIVOT_NUM_FMT

    gt2 = r20 + 1 + len(regions)
    ws.cell(gt2, 1, "Grand Total").font = Font(bold=True)
    running = 0.0
    for ci in range(len(cumm_dates)):
        if ci == 0:
            running = grand[0]
        elif ci < len(delta_dates):
            running += grand[ci]
        c = ws.cell(gt2, 2 + ci, running)
        c.number_format = _PIVOT_NUM_FMT
        c.font = Font(bold=True)

    ws.column_dimensions['A'].width = 29.5
    ws.column_dimensions['B'].width = 15.125
    ws.column_dimensions['H'].width = 15.125
    ws.column_dimensions['I'].width = 15.125


def _create_pivot_tables_fallback(path, rows, singles_rows, dates):
    """Create fallback (non-COM) pivot sheets for environments without Excel COM."""
    wb = openpyxl.load_workbook(path)
    _write_one_pivot_fallback(wb['Pivot Packs'], rows, dates, 'PACKS')
    _write_one_pivot_fallback(wb['Pivot Singles'], singles_rows, dates, 'SINGLES')
    wb.save(path)
    wb.close()


def _create_pivot_tables(path, rows, singles_rows, dates, m0_cycle):
    """Open workbook via COM and create PivotTables for both Packs and Singles."""
    try:
        import win32com.client as win32
        import pythoncom
    except ModuleNotFoundError:
        _create_pivot_tables_fallback(path, rows, singles_rows, dates)
        return

    abs_path = os.path.abspath(path)

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        wb = excel.Workbooks.Open(abs_path)

        _build_one_pivot(wb, 'PivotData', 'Pivot Packs',
                         'PACKS', 'PivotDelta', rows, dates)
        _build_one_pivot(wb, 'PivotData Singles', 'Pivot Singles',
                         'SINGLES', 'PivotDeltaSingles', singles_rows, dates)

        wb.Save()
        wb.Close(False)
    except Exception:
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
        _create_pivot_tables_fallback(path, rows, singles_rows, dates)
        return
    finally:
        if excel:
            try:
                excel.ScreenUpdating = True
                excel.DisplayAlerts = True
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ================================================================
# Streamlit UI
# ================================================================

_MONTH_NAMES = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

_ACTUALS_COL_MAP = {
    "202603": "M", "202604": "N", "202605": "O", "202606": "P",
}


def _parse_prev_filename(name):
    """Extract M-1 cycle from previous compare filename and derive defaults.
    E.g. '202602 vs 202601 IBP Forecast Compare.xlsx' → m1='202602', m0='202603'
    """
    m = re.match(r"(\d{6})\s+vs\s+\d{6}", name)
    if not m:
        return None
    m1_cycle = m.group(1)
    m1_year = int(m1_cycle[:4])
    m1_month = int(m1_cycle[4:6])
    m0_dt = datetime(m1_year, m1_month, 1) + relativedelta(months=1)
    m0_cycle = m0_dt.strftime("%Y%m")
    m0_label = f"{_MONTH_NAMES[m0_dt.month]} Forecast"
    m1_label = f"{_MONTH_NAMES[m1_month]} Forecast"
    actuals_col = _ACTUALS_COL_MAP.get(m0_cycle, "M")
    return m0_cycle, m1_cycle, m0_label, m1_label, actuals_col


def _init_setting_state(prev_file_name):
    """Initialize editable setting fields without overwriting manual changes."""
    defaults = _parse_prev_filename(prev_file_name) if prev_file_name else None

    current_source = st.session_state.get("settings_source_file")
    if current_source == prev_file_name:
        return

    def_m0 = defaults[0] if defaults else "202603"
    def_m1 = defaults[1] if defaults else "202602"
    def_m0_label = defaults[2] if defaults else "Mar Forecast"
    def_m1_label = defaults[3] if defaults else "Feb Forecast"
    def_actuals = defaults[4] if defaults else "M"

    st.session_state["m0_cycle"] = def_m0
    st.session_state["m1_cycle"] = def_m1
    st.session_state["m0_label"] = def_m0_label
    st.session_state["m1_label"] = def_m1_label
    st.session_state["actuals_col"] = def_actuals
    st.session_state["settings_source_file"] = prev_file_name


def _excel_com_available():
    """Return True only when local Excel COM automation is available."""
    try:
        import win32com.client  # noqa: F401
        return True
    except Exception:
        return False


def main():
    st.set_page_config(page_title="IBP Forecast Compare Tool", layout="wide")
    st.title("IBP Forecast Compare Tool")

    com_ok = _excel_com_available()
    if not com_ok:
        st.error(
            "Pivot drill-down 需要本机 Windows + Excel（win32com）。当前环境不支持，"
            "请在本地电脑运行，不要使用云端。"
        )

    # --- File uploaders ---
    st.subheader("Upload Files")
    c1, c2, c3 = st.columns(3)
    ahmed_file = c1.file_uploader("Ahmed File (.xlsb)", type=["xlsb"])
    prev_file = c2.file_uploader("Previous Compare File (.xlsx)", type=["xlsx"])
    master_file = c3.file_uploader("Master Vlookup (.xlsx)", type=["xlsx"])

    # --- Auto-detect cycles from prev filename without clobbering edits ---
    prev_file_name = prev_file.name if prev_file is not None else None
    _init_setting_state(prev_file_name)

    # --- Settings ---
    st.subheader("Settings")
    st.markdown("""
| Cycle | Actuals Month | Column Letter |
|---|---|---|
| 202603 March | 26-Feb | M |
| 202604 April | 26-Mar | N |
| 202605 May | 26-Apr | O |
| 202606 June | 26-May | P |
""")

    s1, s2, s3 = st.columns(3)
    m0_cycle = s1.text_input("M0 Cycle", key="m0_cycle")
    m1_cycle = s2.text_input("M-1 Cycle", key="m1_cycle")
    actuals_col = s3.text_input("Actuals Column Letter", key="actuals_col")

    s4, s5 = st.columns(2)
    m0_label = s4.text_input("M0 Label", key="m0_label")
    m1_label = s5.text_input("M-1 Label", key="m1_label")

    # --- Generate ---
    all_uploaded = ahmed_file is not None and prev_file is not None and master_file is not None
    can_generate = all_uploaded and com_ok
    if st.button("Generate Compare File", type="primary", disabled=not can_generate):
        # pyxlsb needs a file on disk for .xlsb
        tmp_dir = tempfile.mkdtemp()
        ahmed_tmp = os.path.join(tmp_dir, ahmed_file.name)
        with open(ahmed_tmp, "wb") as f:
            f.write(ahmed_file.getvalue())

        with st.spinner("Processing..."):
            result_bytes, count = process(
                ahmed_tmp,
                BytesIO(prev_file.getvalue()),
                BytesIO(master_file.getvalue()),
                m0_cycle, m1_cycle, m0_label, m1_label, actuals_col
            )

        os.remove(ahmed_tmp)
        os.rmdir(tmp_dir)

        out_name = f"{m0_cycle} vs {m1_cycle} IBP Forecast Compare.xlsx"
        st.success(f"Generated: {out_name}  ({count} SKU+Region combinations)")
        st.download_button(
            label="Download Compare File",
            data=result_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    elif not all_uploaded:
        st.info("Please upload all 3 files to enable generation.")
    elif not com_ok:
        st.warning("当前环境不支持 Excel COM，已禁用生成按钮。请在本机 Windows + Excel 环境运行。")


if __name__ == "__main__":
    main()
